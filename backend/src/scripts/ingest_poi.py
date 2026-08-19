"""지도 위 주변시설(POI) 적재 — 학교·지하철역을 공공 표준데이터에서 뽑아 JSON 으로.

네이버 부동산의 **학군**·**교통** 레이어에 해당한다. 두 데이터 모두 실거래 API 처럼
매번 호출하는 방식이 아니라 **파일로 배포되는 표준데이터**라, 한 번 적재해 두고 쓴다.

입력(기본값은 ``data/reference/source/`` 에 둔 원본):
  · 전국초중등학교위치표준데이터.csv          (CP949, 학교ID·학교명·학교급·위도·경도…)
  · 전체_도시철도역사정보_YYYYMMDD.xlsx       (역사명·노선명·환승·역위도·역경도…)

출력(``data/reference/``):
  · schools.json   [{name, level, kind, lat, lng, addr}]
  · stations.json  [{name, line, transfer, lat, lng, addr}]

좌표가 없거나 폐교/미운영인 행은 버린다 — 지도에 못 찍고 찍어도 틀린 정보라서.

사용:
    python -m scripts.ingest_poi
    python -m scripts.ingest_poi --schools ~/Downloads/전국초중등학교위치표준데이터.csv
"""
from __future__ import annotations

import argparse
import csv
import json
from pathlib import Path

from app.core.config import get_settings

# 학교급 → 지도 배지에 쓸 짧은 글자
LEVEL_SHORT = {"초등학교": "초", "중학교": "중", "고등학교": "고", "특수학교": "특"}


def _ref_dir() -> Path:
    d = get_settings().data_dir / "reference"
    d.mkdir(parents=True, exist_ok=True)
    return d


def _src_dir() -> Path:
    return _ref_dir() / "source"


def _num(v) -> float | None:
    try:
        f = float(str(v).strip())
    except (TypeError, ValueError):
        return None
    return f if f else None


def ingest_schools(path: Path) -> list[dict]:
    """학교 표준데이터 CSV(CP949) → 좌표 있는 운영 중 학교만."""
    out: list[dict] = []
    # 표준데이터는 CP949 로 내려온다. 드물게 깨진 바이트가 있어 replace 로 넘긴다.
    with path.open("r", encoding="cp949", errors="replace", newline="") as f:
        for row in csv.DictReader(f):
            if (row.get("운영상태") or "").strip() != "운영":
                continue
            lat, lng = _num(row.get("위도")), _num(row.get("경도"))
            if lat is None or lng is None:
                continue
            level = (row.get("학교급구분") or "").strip()
            out.append({
                "name": (row.get("학교명") or "").strip(),
                "level": level,
                "kind": LEVEL_SHORT.get(level, level[:1] or "학"),
                "lat": round(lat, 6),
                "lng": round(lng, 6),
                "addr": (row.get("소재지도로명주소") or row.get("소재지지번주소") or "").strip(),
            })
    return out


def ingest_stations(path: Path) -> list[dict]:
    """도시철도 역사정보 XLSX → 좌표 있는 역. 같은 역·같은 노선 중복은 하나로."""
    import openpyxl

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = ws.iter_rows(values_only=True)
    header = [str(h).strip() if h is not None else "" for h in next(rows)]
    idx = {h: i for i, h in enumerate(header)}

    def cell(r, key):
        i = idx.get(key)
        return r[i] if i is not None and i < len(r) else None

    seen: set[tuple[str, str]] = set()
    out: list[dict] = []
    for r in rows:
        if r is None:
            continue
        lat, lng = _num(cell(r, "역위도")), _num(cell(r, "역경도"))
        if lat is None or lng is None:
            continue
        name = str(cell(r, "역사명") or "").strip()
        line = str(cell(r, "노선명") or "").strip()
        if not name:
            continue
        key = (name, line)
        if key in seen:
            continue
        seen.add(key)
        transfer = "환승" in str(cell(r, "환승역구분") or "")
        out.append({
            "name": name,
            "line": line,
            "transfer": transfer,
            "lat": round(lat, 6),
            "lng": round(lng, 6),
            "addr": str(cell(r, "역사도로명주소") or "").strip(),
        })
    wb.close()
    return out


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--schools", default=None, help="학교 표준데이터 CSV 경로")
    ap.add_argument("--stations", default=None, help="도시철도 역사정보 XLSX 경로")
    args = ap.parse_args()

    ref, src = _ref_dir(), _src_dir()

    sc_path = Path(args.schools) if args.schools else next(
        iter(sorted(src.glob("전국초중등학교위치표준데이터*.csv"))), None)
    st_path = Path(args.stations) if args.stations else next(
        iter(sorted(src.glob("*도시철도역사정보*.xlsx"), reverse=True)), None)

    if sc_path and sc_path.exists():
        schools = ingest_schools(sc_path)
        (ref / "schools.json").write_text(
            json.dumps(schools, ensure_ascii=False, separators=(",", ":")), encoding="utf-8")
        by_level: dict[str, int] = {}
        for s in schools:
            by_level[s["level"]] = by_level.get(s["level"], 0) + 1
        print(f"학교   {len(schools):,}곳 → {ref / 'schools.json'}")
        for k, v in sorted(by_level.items(), key=lambda x: -x[1]):
            print(f"        {k} {v:,}")
    else:
        print(f"학교 CSV 를 찾지 못했습니다 — {src} 에 넣거나 --schools 로 지정하세요.")

    if st_path and st_path.exists():
        stations = ingest_stations(st_path)
        (ref / "stations.json").write_text(
            json.dumps(stations, ensure_ascii=False, separators=(",", ":")), encoding="utf-8")
        lines = len({s["line"] for s in stations})
        print(f"지하철 {len(stations):,}역(노선 {lines}개) → {ref / 'stations.json'}")
    else:
        print(f"도시철도 XLSX 를 찾지 못했습니다 — {src} 에 넣거나 --stations 로 지정하세요.")


if __name__ == "__main__":
    main()
